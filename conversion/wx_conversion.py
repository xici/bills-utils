#!/usr/bin/env python3

import datetime
import os
import sys
import argparse
import pathlib
import pdfplumber
import openpyxl


def convert_pdf(pdf_file, output_dir=None):
    """
    转换PDF文件为Excel

    Args:
        pdf_file: PDF文件路径
        output_dir: 输出目录（可选，默认为PDF文件所在目录）

    Returns:
        str: 生成的Excel文件路径，失败时返回None
    """
    try:
        if not os.path.exists(pdf_file):
            print(f"错误: 文件不存在: {pdf_file}")
            return None

        if not pdf_file.lower().endswith(".pdf"):
            print(f"错误: 文件不是PDF格式: {pdf_file}")
            return None

        # 确定输出目录
        if output_dir is None:
            output_dir = os.path.dirname(os.path.abspath(pdf_file))

        # 确保输出目录存在
        os.makedirs(output_dir, exist_ok=True)

        print(f"正在处理文件: {pdf_file}")
        print(f"输出目录: {output_dir}")

        total_rows = []
        with pdfplumber.open(pdf_file) as pdf:
            for page in pdf.pages:
                tables = page.extract_tables()
                for table in tables:
                    for row in table:
                        if len(row) > 1:
                            # 清理数据
                            if row[1] and "\n" in row[1]:
                                row[1] = row[1].replace("\n", " ")
                            row = [col.replace("\n", "") if col else "" for col in row]
                            total_rows.append(row)

        if len(total_rows) < 4:
            print("警告: 文件中未找到表格")
            return None

        # 提取表头和数据
        header = total_rows[2]
        data_rows = total_rows[3:]

        # 按日期排序
        data_rows.sort(key=lambda ele: ele[1])

        # 创建Excel文件
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "账单"

        # 写入表头
        for ci, col in enumerate(header):
            sheet.cell(row=1, column=ci + 1, value=col)

        # 写入数据
        for ri, row in enumerate(data_rows):
            for ci, col in enumerate(row):
                if ci == 1:  # 日期转格式
                    try:
                        sheet.cell(
                            row=ri + 2,
                            column=ci + 1,
                            value=datetime.datetime.fromisoformat(col),
                        )
                        sheet.cell(row=ri + 2, column=ci + 1).number_format = (
                            "yyyy-mm-dd hh:mm:ss"
                        )
                    except:
                        sheet.cell(row=ri + 2, column=ci + 1, value=col)
                elif ci == 5:  # 金额转小数
                    try:
                        sheet.cell(row=ri + 2, column=ci + 1, value=float(col))
                        sheet.cell(row=ri + 2, column=ci + 1).number_format = (
                            "0.00"  # 保留两位小数
                        )
                    except:
                        sheet.cell(row=ri + 2, column=ci + 1, value=col)
                else:
                    sheet.cell(row=ri + 2, column=ci + 1, value=col)

        # 生成输出文件名
        pdf_name = os.path.splitext(os.path.basename(pdf_file))[0]
        timestamp = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
        output_file = os.path.join(output_dir, f"{pdf_name}_converted_{timestamp}.xlsx")

        # 保存文件
        wb.save(output_file)

        print(f"转换成功!")
        print(f"  共处理 {len(data_rows)} 条记录")
        print(f"  输出文件: {output_file}")

        return output_file

    except Exception as e:
        print(f"转换失败: {e}")
        import traceback

        traceback.print_exc()
        return None


def convert_directory(pdf_dir, output_dir=None):
    """
    转换目录中的所有PDF文件

    Args:
        pdf_dir: 包含PDF文件的目录
        output_dir: 输出目录（可选）

    Returns:
        list: 成功转换的文件列表
    """
    if not os.path.isdir(pdf_dir):
        print(f"错误: 目录不存在: {pdf_dir}")
        return []

    # 查找所有PDF文件
    pdf_files = list(pathlib.Path(pdf_dir).glob("*.pdf"))

    if not pdf_files:
        print(f"警告: 目录中未找到PDF文件: {pdf_dir}")
        return []

    print(f"在目录中找到 {len(pdf_files)} 个PDF文件")

    success_files = []
    for pdf_file in pdf_files:
        print(f"\n处理文件 {pdf_file.name}...")
        result = convert_pdf(str(pdf_file), output_dir)
        if result:
            success_files.append(result)

    return success_files


def main():
    """主函数"""
    parser = argparse.ArgumentParser(
        description="微信账单PDF转Excel工具",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
使用示例:
  # 转换单个PDF文件
  python %(prog)s bill.pdf
  
  # 转换单个PDF文件并指定输出目录
  python %(prog)s bill.pdf ./output
  
  # 转换目录中的所有PDF文件
  python %(prog)s ./pdf_files
  
  # 转换目录中的所有PDF文件并指定输出目录
  python %(prog)s ./pdf_files ./output
        """,
    )

    parser.add_argument("input", help="PDF文件路径或包含PDF文件的目录")

    parser.add_argument(
        "output_dir",
        nargs="?",
        help="输出目录（可选，默认为输入文件所在目录或当前目录）",
    )

    parser.add_argument(
        "-v",
        "--version",
        action="version",
        version="微信账单 PDF 转 Excel 工具",
    )

    args = parser.parse_args()

    input_path = args.input
    output_dir = args.output_dir

    # 检查输入路径是否存在
    if not os.path.exists(input_path):
        print(f"错误: 输入路径不存在: {input_path}")
        return 1

    # 判断输入是文件还是目录
    if os.path.isfile(input_path):
        # 单个文件转换
        result = convert_pdf(input_path, output_dir)
        if result:
            print(f"\n✓ 转换完成: {result}")
            return 0
        else:
            print("\n✗ 转换失败")
            return 1
    else:
        # 目录批量转换
        success_files = convert_directory(input_path, output_dir)

        if success_files:
            print(f"\n✓ 批量转换完成!")
            print(f"  成功转换 {len(success_files)} 个文件:")
            for file in success_files:
                print(f"    - {file}")
            return 0
        else:
            print("\n✗ 批量转换失败或无文件可转换")
            return 1


if __name__ == "__main__":
    sys.exit(main())
